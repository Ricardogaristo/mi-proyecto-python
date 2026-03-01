from flask import Flask, render_template, request, redirect, session, send_file, url_for, jsonify
from functools import wraps
from datetime import datetime
from collections import defaultdict
import io, os, json, threading, time
from dotenv import load_dotenv
load_dotenv()

# ── BD: PostgreSQL en Render, MariaDB en local ─────────────────────────────
DATABASE_URL = os.getenv("DATABASE_URL", "")
_USE_PG = bool(DATABASE_URL)
if _USE_PG:
    import psycopg2, psycopg2.extras
else:
    import pymysql, pymysql.cursors

# Para crear Excel bonito
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

# Módulo de Formación
from formacion import formacion_bp, inicializar_formacion

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "clave_secreta_muy_segura")

# Registrar Blueprint de Formación
app.register_blueprint(formacion_bp)

# --- BASE DE DATOS ---

def get_connection():
    if _USE_PG:
        return psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)
    return pymysql.connect(
        host="localhost", port=3306, db="gestor_tareas",
        user="root", password="", charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor,
    )

def inicializar_todo():
    conn = get_connection()
    with conn.cursor() as cursor:
        if _USE_PG:
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS usuarios (
                    id       SERIAL PRIMARY KEY,
                    username VARCHAR(100) NOT NULL UNIQUE,
                    email    VARCHAR(255) UNIQUE,
                    password VARCHAR(255) NOT NULL,
                    es_admin SMALLINT NOT NULL DEFAULT 0
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS tareas (
                    id          SERIAL PRIMARY KEY,
                    descripcion TEXT NOT NULL,
                    categoria   VARCHAR(150),
                    fecha       DATE,
                    completada  SMALLINT NOT NULL DEFAULT 0,
                    codigo      VARCHAR(50),
                    usuario_id  INT REFERENCES usuarios(id) ON DELETE SET NULL,
                    prioridad   SMALLINT NOT NULL DEFAULT 2,
                    favorita    SMALLINT NOT NULL DEFAULT 0,
                    notas       TEXT
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS subtareas (
                    id       SERIAL PRIMARY KEY,
                    tarea_id INT NOT NULL REFERENCES tareas(id) ON DELETE CASCADE,
                    texto    TEXT NOT NULL,
                    hecha    SMALLINT NOT NULL DEFAULT 0
                )
            """)
            cursor.execute("""
                INSERT INTO usuarios (username, email, password, es_admin)
                VALUES (%s,%s,%s,%s) ON CONFLICT (username) DO NOTHING
            """, ("admin","admin@correo.com","1234",1))
        else:
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS usuarios (
                    id        SERIAL PRIMARY KEY,
                    username  VARCHAR(100) NOT NULL,
                    email     VARCHAR(255),
                    password  VARCHAR(255) NOT NULL,
                    es_admin  TINYINT(1) NOT NULL DEFAULT 0,
                    PRIMARY KEY (id),
                    UNIQUE KEY uq_username (username),
                    UNIQUE KEY uq_email (email)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS tareas (
                    id          SERIAL PRIMARY KEY,
                    descripcion TEXT NOT NULL,
                    categoria   VARCHAR(150),
                    fecha       DATE,
                    completada  TINYINT(1) NOT NULL DEFAULT 0,
                    codigo      VARCHAR(50),
                    usuario_id  INT,
                    prioridad   TINYINT NOT NULL DEFAULT 2,
                    favorita    TINYINT(1) NOT NULL DEFAULT 0,
                    notas       TEXT,
                    PRIMARY KEY (id),
                    CONSTRAINT fk_tarea_usuario FOREIGN KEY (usuario_id) REFERENCES usuarios(id) ON DELETE SET NULL
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS subtareas (
                    id       SERIAL PRIMARY KEY,
                    tarea_id INT NOT NULL,
                    texto    TEXT NOT NULL,
                    hecha    TINYINT(1) NOT NULL DEFAULT 0,
                    PRIMARY KEY (id),
                    CONSTRAINT fk_subtarea_tarea FOREIGN KEY (tarea_id) REFERENCES tareas(id) ON DELETE CASCADE
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
            cursor.execute("""
                INSERT IGNORE INTO usuarios (username, email, password, es_admin)
                VALUES (%s,%s,%s,%s)
            """, ("admin","admin@correo.com","1234",1))
    conn.commit()
    conn.close()
    print("✅ Base de datos inicializada.")


# --- PROTECCIÓN DE RUTAS ---
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        if session.get("es_admin") != 1:
            return redirect("/")
        return f(*args, **kwargs)
    return decorated_function

# --- RUTAS ---
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        identificador = request.form.get("username")
        password = request.form.get("password")

        conn = get_connection()
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT id, username, es_admin FROM usuarios
                WHERE (username=%s OR email=%s) AND password=%s
            """, (identificador, identificador, password))
            usuario = cursor.fetchone()
        conn.close()

        if usuario:
            session["user_id"]  = usuario["id"]
            session["user"]     = usuario["username"]
            session["es_admin"] = usuario["es_admin"]
            return redirect("/")

    return render_template("login.html", error="Credenciales incorrectas")


@app.route("/accesos_rapidos")
@login_required
def accesos_rapidos():
    return render_template("accesos_rapidos.html")


@app.route("/registro", methods=["GET", "POST"])
def registro():
    if request.method == "POST":
        username = request.form.get("username")
        email    = request.form.get("email")
        password = request.form.get("password")

        conn = get_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute("""
                    INSERT INTO usuarios (username, email, password, es_admin)
                    VALUES (%s, %s, %s, 0)
                """, (username, email, password))
            conn.commit()
            conn.close()
            return redirect(url_for("login"))
        except Exception as e:
            conn.close()
            if "unique" in str(e).lower() or "duplicate" in str(e).lower() or "1062" in str(e):
                return render_template("registro.html", error="El usuario o email ya existe ❌")
            raise

    return render_template("registro.html")


@app.route("/")
@login_required
def index():
    user_id  = session.get("user_id")
    es_admin = session.get("es_admin")

    filtro_estado = request.args.get("estado", "all").strip()
    filtro_cat    = request.args.get("cat",    "").strip()
    filtro_q      = request.args.get("q",      "").strip()
    filtro_prio   = request.args.get("prio",   "").strip()
    filtro_fav    = request.args.get("fav",    "").strip()
    page     = max(request.args.get("page", 1, type=int), 1)
    per_page = 10

    filtros = []
    params  = []

    if es_admin != 1:
        filtros.append("usuario_id = %s")
        params.append(user_id)

    if filtro_estado == "pending":
        filtros.append("completada = 0")
    elif filtro_estado == "done":
        filtros.append("completada = 1")

    if filtro_cat:
        filtros.append("LOWER(TRIM(COALESCE(categoria,''))) = LOWER(TRIM(%s))")
        params.append(filtro_cat)

    if filtro_prio in ("1","2","3"):
        filtros.append("prioridad = %s")
        params.append(int(filtro_prio))

    if filtro_fav == "1":
        filtros.append("favorita = 1")

    if filtro_q:
        filtros.append("(LOWER(descripcion) LIKE %s OR LOWER(COALESCE(codigo,'')) LIKE %s OR LOWER(COALESCE(categoria,'')) LIKE %s)")
        like = f"%{filtro_q.lower()}%"
        params.extend([like, like, like])

    where_clause = ("WHERE " + " AND ".join(filtros)) if filtros else ""

    conn   = get_connection()
    with conn.cursor() as cursor:
        cursor.execute(f"SELECT COUNT(*) AS cnt FROM tareas {where_clause}", params)
        total_tareas = cursor.fetchone()["cnt"]

        total_pages = max((total_tareas + per_page - 1) // per_page, 1)
        if page > total_pages:
            page = total_pages
        offset = (page - 1) * per_page

        cursor.execute(
            f"SELECT * FROM tareas {where_clause} ORDER BY favorita DESC, prioridad ASC, id DESC LIMIT %s OFFSET %s",
            params + [per_page, offset]
        )
        tareas_raw = cursor.fetchall()

        # index.html accede por índice (t[0], t[1]…)
        # Orden: id, descripcion, categoria, fecha, completada, codigo, usuario_id, prioridad, favorita, notas
        tareas = [
            (
                t["id"],
                t["descripcion"],
                t["categoria"],
                t["fecha"].strftime("%Y-%m-%d") if t["fecha"] else "",
                t["completada"],
                t["codigo"],
                t["usuario_id"],
                t["prioridad"] or 2,
                t["favorita"] or 0,
                t["notas"] or ""
            )
            for t in tareas_raw
        ]

        ids = [t[0] for t in tareas]
        subtareas_map = {}
        if ids:
            fmt = ",".join(["%s"] * len(ids))
            cursor.execute(f"SELECT * FROM subtareas WHERE tarea_id IN ({fmt}) ORDER BY id", ids)
            for s in cursor.fetchall():
                subtareas_map.setdefault(s["tarea_id"], []).append(dict(s))

        if es_admin == 1:
            cursor.execute("""
                SELECT DISTINCT COALESCE(NULLIF(TRIM(categoria),''), 'General') AS cat
                FROM tareas ORDER BY cat
            """)
        else:
            cursor.execute("""
                SELECT DISTINCT COALESCE(NULLIF(TRIM(categoria),''), 'General') AS cat
                FROM tareas WHERE usuario_id = %s ORDER BY cat
            """, (user_id,))
        categorias_lista = [row["cat"] for row in cursor.fetchall()]

    conn.close()

    return render_template(
        "index.html",
        tareas=tareas,
        page=page,
        total_pages=total_pages,
        filtro_estado=filtro_estado,
        filtro_cat=filtro_cat,
        filtro_q=filtro_q,
        filtro_prio=filtro_prio,
        filtro_fav=filtro_fav,
        categorias_lista=categorias_lista,
        subtareas_map=subtareas_map,
    )


@app.route("/agregar", methods=["POST"])
@login_required
def agregar():
    codigo      = request.form.get("codigo")
    descripcion = request.form.get("descripcion")
    categoria   = request.form.get("categoria")
    fecha       = request.form.get("fecha") or None
    prioridad   = int(request.form.get("prioridad", 2))
    notas       = request.form.get("notas", "")

    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute("""
            INSERT INTO tareas (codigo, descripcion, categoria, fecha,
                                completada, usuario_id, prioridad, favorita, notas)
            VALUES (%s, %s, %s, %s, 0, %s, %s, 0, %s)
        """, (codigo or None, descripcion, categoria or None, fecha,
              session["user_id"], prioridad, notas or None))
    conn.commit()
    conn.close()
    return redirect("/")


@app.route("/admin")
@admin_required
def admin():
    filtro_cat = request.args.get("categoria", "").strip()
    filtro_est = request.args.get("estado",    "").strip()
    page       = max(request.args.get("page", 1, type=int), 1)
    per_page   = 10

    filtros = []
    params  = []

    if filtro_cat:
        filtros.append("LOWER(TRIM(categoria)) = LOWER(TRIM(%s))")
        params.append(filtro_cat)
    if filtro_est == "Completada":
        filtros.append("completada = 1")
    elif filtro_est == "Pendiente":
        filtros.append("completada = 0")

    where_clause = ("WHERE " + " AND ".join(filtros)) if filtros else ""

    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute(f"SELECT COUNT(*) AS cnt FROM tareas {where_clause}", params)
        total_filtrado = cursor.fetchone()["cnt"]
        total_pages = max((total_filtrado + per_page - 1) // per_page, 1)
        if page > total_pages:
            page = total_pages
        offset = (page - 1) * per_page

        cursor.execute(f"""
            SELECT id, descripcion, categoria, fecha,
                   completada, codigo, usuario_id
            FROM tareas {where_clause}
            ORDER BY id DESC
            LIMIT %s OFFSET %s
        """, params + [per_page, offset])
        tareas = cursor.fetchall()

        cursor.execute("""
            SELECT DISTINCT categoria FROM tareas
            WHERE categoria IS NOT NULL AND categoria != ''
            ORDER BY categoria
        """)
        categorias_lista = [row["categoria"] for row in cursor.fetchall()]
    conn.close()

    # Convertir tareas a tuplas para compatibilidad con template existente
    tareas_tuple = [
        (
            t["id"], t["descripcion"], t["categoria"],
            t["fecha"].strftime("%Y-%m-%d") if t["fecha"] else "",
            t["completada"], t["codigo"], t["usuario_id"]
        )
        for t in tareas
    ]

    return render_template(
        "admin.html",
        tareas=tareas_tuple,
        page=page,
        total_pages=total_pages,
        total=total_filtrado,
        categorias=categorias_lista,
        filtro_cat=filtro_cat,
        filtro_est=filtro_est,
    )


# ══════════════════════════════════════════════════════
# GESTIÓN DE USUARIOS (solo admin)
# ══════════════════════════════════════════════════════

@app.route("/usuarios")
@admin_required
def usuarios():
    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute("""
            SELECT u.id, u.username, u.email, u.es_admin,
                   COUNT(t.id)                                      AS total_tareas,
                   SUM(CASE WHEN t.completada=1 THEN 1 ELSE 0 END) AS completadas,
                   SUM(CASE WHEN t.completada=0 THEN 1 ELSE 0 END) AS pendientes
            FROM usuarios u
            LEFT JOIN tareas t ON t.usuario_id = u.id
            GROUP BY u.id, u.username, u.email, u.es_admin
            ORDER BY u.es_admin DESC, u.username
        """)
        usuarios_lista = cursor.fetchall()

        cursor.execute("""
            SELECT DISTINCT COALESCE(NULLIF(TRIM(categoria),''),'General') AS cat
            FROM tareas ORDER BY cat
        """)
        categorias = [r["cat"] for r in cursor.fetchall()]
    conn.close()
    return render_template("usuarios.html",
                           usuarios=usuarios_lista,
                           categorias=categorias)


@app.route("/usuarios/eliminar/<int:uid>", methods=["POST"])
@admin_required
def usuario_eliminar(uid):
    if uid == session["user_id"]:
        return redirect("/usuarios")
    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute("DELETE FROM tareas   WHERE usuario_id = %s", (uid,))
        cursor.execute("DELETE FROM usuarios WHERE id = %s",         (uid,))
    conn.commit()
    conn.close()
    return redirect("/usuarios")


@app.route("/usuarios/asignar_tarea/<int:uid>", methods=["POST"])
@admin_required
def usuario_asignar_tarea(uid):
    descripcion = request.form.get("descripcion", "").strip()
    if not descripcion:
        return redirect("/usuarios")
    codigo    = request.form.get("codigo",    "").strip() or None
    categoria = request.form.get("categoria", "").strip() or None
    fecha     = request.form.get("fecha",     "").strip() or None
    prioridad = int(request.form.get("prioridad", 2))
    notas     = request.form.get("notas",     "").strip() or None

    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute("""
            INSERT INTO tareas (codigo, descripcion, categoria, fecha,
                                completada, usuario_id, prioridad, favorita, notas)
            VALUES (%s,%s,%s,%s,0,%s,%s,0,%s)
        """, (codigo, descripcion, categoria, fecha, uid, prioridad, notas))
    conn.commit()
    conn.close()
    return redirect("/usuarios")


@app.route("/usuarios/toggle_admin/<int:uid>", methods=["POST"])
@admin_required
def usuario_toggle_admin(uid):
    if uid == session["user_id"]:
        return redirect("/usuarios")
    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute("UPDATE usuarios SET es_admin = 1 - es_admin WHERE id=%s", (uid,))
    conn.commit()
    conn.close()
    return redirect("/usuarios")


@app.route("/completar/<int:id>")
@login_required
def completar(id):
    conn = get_connection()
    with conn.cursor() as cursor:
        if session.get("es_admin") == 1:
            cursor.execute("UPDATE tareas SET completada=1 WHERE id=%s", (id,))
        else:
            cursor.execute("UPDATE tareas SET completada=1 WHERE id=%s AND usuario_id=%s",
                           (id, session["user_id"]))
    conn.commit()
    conn.close()
    return redirect("/")


@app.route("/eliminar/<int:id>")
@login_required
def eliminar(id):
    conn = get_connection()
    with conn.cursor() as cursor:
        if session["es_admin"] == 1:
            cursor.execute("DELETE FROM tareas WHERE id=%s", (id,))
        else:
            cursor.execute("DELETE FROM tareas WHERE id=%s AND usuario_id=%s",
                           (id, session["user_id"]))
    conn.commit()
    conn.close()
    return redirect("/")


@app.route("/dashboard")
@login_required
def dashboard():
    user_id  = session.get("user_id")
    es_admin = session.get("es_admin")

    if es_admin == 1:
        where  = ""
        params = ()
    else:
        where  = "WHERE usuario_id = %s"
        params = (user_id,)

    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute(f"""
            SELECT COUNT(*) AS total,
                   SUM(CASE WHEN completada=1 THEN 1 ELSE 0 END) AS completadas
            FROM tareas {where}
        """, params)
        stats       = cursor.fetchone()
        total       = stats["total"]       or 0
        completadas = stats["completadas"] or 0
        pendientes  = total - completadas
        porcentaje  = round((completadas / total) * 100, 1) if total > 0 else 0

        if porcentaje > 70:
            nivel, color_nivel = "Alto",  "success"
        elif porcentaje > 40:
            nivel, color_nivel = "Medio", "warning"
        else:
            nivel, color_nivel = "Bajo",  "danger"

        cursor.execute(f"SELECT * FROM tareas {where} ORDER BY id DESC LIMIT 5", params)
        ultimas_tareas = cursor.fetchall()

        cursor.execute(f"""
            SELECT categoria, COUNT(*) AS cantidad
            FROM tareas {where}
            GROUP BY categoria ORDER BY cantidad DESC
        """, params)
        datos_categorias = cursor.fetchall()
    conn.close()

    categorias = [f["categoria"] if f["categoria"] else "General" for f in datos_categorias]
    cantidades = [f["cantidad"] for f in datos_categorias]
    fecha_actual = datetime.now().strftime("%d/%m/%Y")

    return render_template("dashboard.html",
                           total=total,
                           completadas=completadas,
                           pendientes=pendientes,
                           porcentaje=porcentaje,
                           nivel=nivel,
                           color_nivel=color_nivel,
                           fecha_actual=fecha_actual,
                           ultimas_tareas=ultimas_tareas,
                           categorias=categorias,
                           cantidades=cantidades)


# ── Favorita toggle ──────────────────────────────────────────────────────
@app.route("/favorita/<int:id>")
@login_required
def favorita(id):
    conn = get_connection()
    with conn.cursor() as cursor:
        if session.get("es_admin") == 1:
            cursor.execute("UPDATE tareas SET favorita = 1 - favorita WHERE id=%s", (id,))
        else:
            cursor.execute("UPDATE tareas SET favorita = 1 - favorita WHERE id=%s AND usuario_id=%s",
                           (id, session["user_id"]))
    conn.commit()
    conn.close()
    return redirect(request.referrer or "/")


# ── Duplicar tarea ───────────────────────────────────────────────────────
@app.route("/duplicar/<int:id>")
@login_required
def duplicar(id):
    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute("SELECT * FROM tareas WHERE id=%s", (id,))
        t = cursor.fetchone()
        if t:
            cursor.execute("""
                INSERT INTO tareas (descripcion, categoria, fecha, completada,
                                    codigo, usuario_id, prioridad, favorita, notas)
                VALUES (%s,%s,%s,0,%s,%s,%s,0,%s)
            """, (
                t["descripcion"] + " (copia)", t["categoria"], t["fecha"],
                t["codigo"], t["usuario_id"], t["prioridad"] or 2, t["notas"] or ""
            ))
    conn.commit()
    conn.close()
    return redirect(request.referrer or "/")


# ── Subtareas: agregar ───────────────────────────────────────────────────
@app.route("/subtarea/agregar/<int:tarea_id>", methods=["POST"])
@login_required
def subtarea_agregar(tarea_id):
    texto = request.form.get("texto", "").strip()
    if texto:
        conn = get_connection()
        with conn.cursor() as cursor:
            cursor.execute("INSERT INTO subtareas (tarea_id, texto) VALUES (%s,%s)", (tarea_id, texto))
        conn.commit()
        conn.close()
    return redirect(request.referrer or "/")


# ── Subtareas: toggle hecha ──────────────────────────────────────────────
@app.route("/subtarea/toggle/<int:sub_id>")
@login_required
def subtarea_toggle(sub_id):
    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute("UPDATE subtareas SET hecha = 1 - hecha WHERE id=%s", (sub_id,))
    conn.commit()
    conn.close()
    return redirect(request.referrer or "/")


# ── Subtareas: eliminar ──────────────────────────────────────────────────
@app.route("/subtarea/eliminar/<int:sub_id>")
@login_required
def subtarea_eliminar(sub_id):
    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute("DELETE FROM subtareas WHERE id=%s", (sub_id,))
    conn.commit()
    conn.close()
    return redirect(request.referrer or "/")


@app.route("/editar/<int:id>", methods=["GET", "POST"])
@login_required
def editar(id):
    conn = get_connection()
    if request.method == "POST":
        codigo      = request.form.get("codigo")      or None
        descripcion = request.form.get("descripcion")
        categoria   = request.form.get("categoria")   or None
        fecha       = request.form.get("fecha")       or None
        completada  = request.form.get("completada")
        prioridad   = int(request.form.get("prioridad", 2))
        notas       = request.form.get("notas", "") or None

        with conn.cursor() as cursor:
            cursor.execute("""
                UPDATE tareas
                SET codigo=%s, descripcion=%s, categoria=%s, fecha=%s,
                    completada=%s, prioridad=%s, notas=%s
                WHERE id=%s
            """, (codigo, descripcion, categoria, fecha, completada, prioridad, notas, id))
        conn.commit()
        conn.close()
        return redirect("/")
    else:
        with conn.cursor() as cursor:
            cursor.execute("SELECT * FROM tareas WHERE id=%s", (id,))
            tarea_dict = cursor.fetchone()
        conn.close()
        # Convertir a tupla para compatibilidad con el template existente
        # tarea[0]=id  [1]=descripcion  [2]=categoria  [3]=fecha  [4]=completada  [5]=codigo
        if tarea_dict:
            # MariaDB devuelve fecha como objeto date, convertir a string
            fecha_str = tarea_dict["fecha"].strftime("%Y-%m-%d") if tarea_dict["fecha"] else ""
            tarea = (
                tarea_dict["id"],
                tarea_dict["descripcion"],
                tarea_dict["categoria"],
                fecha_str,
                tarea_dict["completada"],
                tarea_dict["codigo"],
            )
        else:
            tarea = None
        return render_template("editar.html", tarea=tarea)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ─────────────────────────────────────────────────────────────────────────────
# EXPORTAR EXCEL
# ─────────────────────────────────────────────────────────────────────────────

_DARK   = "1A1410"; _DARK2  = "241C16"; _ACCENT = "E6A15A"; _AMBER  = "D18B47"
_BLUE   = "C47A5A"; _PURPLE = "B07A63"; _TEAL   = "8F6B50"
_TXT_W  = "F5EDE6"; _TXT_M  = "B7A79A"; _BDR    = "3A2E25"
_CAT_COLS = [
    ("2A1F18","E6A15A"),("332419","D18B47"),("2E1E17","C47A5A"),
    ("2B1B16","B07A63"),("241A15","8F6B50"),("3A221C","E07A5F"),
    ("2F2018","F2A65A"),("352317","C97B3C"),
]

def _xb():
    s = Side(style="thin", color=_BDR)
    return Border(left=s, right=s, top=s, bottom=s)
def _xf(h):  return PatternFill(start_color=h, end_color=h, fill_type="solid")
def _xfn(bold=False, color=_TXT_W, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def _xal(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _xhdr(cell, bg=_DARK, fg=_TXT_W, size=10):
    cell.fill = _xf(bg); cell.font = _xfn(bold=True, color=fg, size=size)
    cell.alignment = _xal(); cell.border = _xb()
def _xaw(ws, min_w=8, max_w=50):
    for col_cells in ws.iter_cols():
        first = col_cells[0]
        if isinstance(first, MergedCell): continue
        length = max(
            (len(str(c.value)) if c.value is not None and not isinstance(c, MergedCell) else 0)
            for c in col_cells
        )
        ws.column_dimensions[first.column_letter].width = min(max(length + 3, min_w), max_w)


def _hoja_resumen(wb, todas, hoy_str, es_admin_usr, username):
    ws = wb.active; ws.title = "Resumen General"
    ws.sheet_properties.tabColor = _ACCENT; ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1"); c = ws["A1"]
    c.value = f"  REPORTE DE TAREAS — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.fill = _xf(_DARK); c.font = _xfn(bold=True, size=14, color=_ACCENT)
    c.alignment = _xal(h="left"); c.border = _xb(); ws.row_dimensions[1].height = 36
    ws.merge_cells("A2:H2"); c = ws["A2"]
    c.value = f"  Generado por: {username}   ·   {'Vista administrador' if es_admin_usr else 'Vista personal'}"
    c.fill = _xf(_DARK2); c.font = _xfn(color=_TXT_M, size=9, italic=True)
    c.alignment = _xal(h="left"); c.border = _xb(); ws.row_dimensions[2].height = 20
    total = len(todas); comp = sum(1 for t in todas if t["completada"]==1)
    pend = total-comp; pct = round(comp/total*100,1) if total else 0
    n_cats = len({t["categoria"] or "General" for t in todas})
    kpis = [("TOTAL TAREAS",total,_BLUE,"0D2040"),("COMPLETADAS",comp,_ACCENT,"1A4A24"),
            ("PENDIENTES",pend,_AMBER,"3D2E00"),(f"% COMPLETADO",f"{pct}%",_TEAL,"0A2E2A"),
            ("CATEGORÍAS",n_cats,_PURPLE,"1E1040")]
    for rh,h in [(3,10),(4,22),(5,40),(6,18),(7,10)]: ws.row_dimensions[rh].height = h
    for ci,(label,valor,fg,bg) in enumerate(kpis,start=1):
        for rn,val,fsz in [(4,label,8),(5,valor,22),(6,"",9)]:
            c = ws.cell(row=rn,column=ci,value=val)
            c.fill=_xf(bg); c.font=_xfn(bold=True,color=fg,size=fsz)
            c.alignment=_xal(); c.border=_xb()
        ws.column_dimensions[get_column_letter(ci)].width = 18
    ws.row_dimensions[8].height = 22
    for ci,h in enumerate(["Categoría","Total","Completadas","Pendientes","% Completado","Usuarios"],start=1):
        c = ws.cell(row=8,column=ci,value=h); _xhdr(c,bg=_DARK,fg=_ACCENT)
    cat_data = defaultdict(lambda:{"total":0,"completadas":0,"usuarios":set()})
    for t in todas:
        cat = t["categoria"] or "General"
        cat_data[cat]["total"] += 1
        cat_data[cat]["completadas"] += (1 if t["completada"]==1 else 0)
        if t.get("username"): cat_data[cat]["usuarios"].add(t["username"])
    rn = 9
    for ci2,(cat,d) in enumerate(sorted(cat_data.items())):
        bg_f,fg_c = _CAT_COLS[ci2 % len(_CAT_COLS)]
        pc2=d["total"]-d["completadas"]; ptc2=round(d["completadas"]/d["total"]*100,1) if d["total"] else 0
        u_str=", ".join(sorted(d["usuarios"])) if d["usuarios"] else "—"
        for cj,val in enumerate([cat,d["total"],d["completadas"],pc2,f"{ptc2}%",u_str],start=1):
            c=ws.cell(row=rn,column=cj,value=val)
            c.fill=_xf(bg_f if cj==1 else _DARK2); c.font=_xfn(bold=(cj==1),color=fg_c if cj==1 else _TXT_W)
            c.alignment=_xal(h="left" if cj in (1,6) else "center"); c.border=_xb()
        ws.row_dimensions[rn].height=18; rn+=1
    for cj,val in enumerate(["TOTAL",total,comp,pend,f"{pct}%",""],start=1):
        c=ws.cell(row=rn,column=cj,value=val)
        c.fill=_xf(_DARK); c.font=_xfn(bold=True,color=_ACCENT,size=10)
        c.alignment=_xal(h="left" if cj==1 else "center"); c.border=_xb()
    ws.row_dimensions[rn].height=20
    for i,w in enumerate([28,10,14,12,16,32],start=1):
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="A9"


def _hoja_categoria(wb, cat_nombre, tareas_cat, color_idx):
    bg_fill,fg_col = _CAT_COLS[color_idx % len(_CAT_COLS)]
    titulo = cat_nombre[:28]+"..." if len(cat_nombre)>31 else cat_nombre
    ws = wb.create_sheet(title=titulo); ws.sheet_properties.tabColor=fg_col; ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:G1"); c=ws["A1"]; c.value=f"  {cat_nombre.upper()}"
    c.fill=_xf(bg_fill); c.font=_xfn(bold=True,size=13,color=fg_col)
    c.alignment=_xal(h="left"); c.border=_xb(); ws.row_dimensions[1].height=32
    tc=len(tareas_cat); cc=sum(1 for t in tareas_cat if t["completada"]==1)
    pc=tc-cc; ptc=round(cc/tc*100,1) if tc else 0
    ws.merge_cells("A2:G2"); c=ws["A2"]
    c.value=f"  Total: {tc}  ·  Completadas: {cc}  ·  Pendientes: {pc}  ·  Progreso: {ptc}%"
    c.fill=_xf(_DARK2); c.font=_xfn(color=_TXT_M,size=9,italic=True)
    c.alignment=_xal(h="left"); c.border=_xb(); ws.row_dimensions[2].height=18
    for ci,h in enumerate(["ID","Código","Descripción","Fecha","Estado","Usuario","Notas"],start=1):
        _xhdr(ws.cell(row=3,column=ci,value=h),bg=bg_fill,fg=fg_col)
    ws.row_dimensions[3].height=20
    for ri,t in enumerate(tareas_cat,start=4):
        done=t["completada"]==1; est="✔  Completada" if done else "●  Pendiente"
        ec=_ACCENT if done else _AMBER; rbg="1A4A24" if done else "3D2E00"
        vals=[t["id"],t["codigo"] or "—",t["descripcion"],t["fecha"] or "—",est,t.get("username") or "—",""]
        for ci,val in enumerate(vals,start=1):
            c=ws.cell(row=ri,column=ci,value=val)
            c.fill=_xf(rbg if ci==5 else _DARK2)
            c.font=_xfn(bold=(ci==5),color=ec if ci==5 else (_TXT_M if ci in (1,4,6) else _TXT_W),size=9 if ci in (1,4,6) else 10)
            c.alignment=_xal(h="center" if ci in (1,4,5) else "left",wrap=(ci==3)); c.border=_xb()
        ws.row_dimensions[ri].height=16
    _xaw(ws); ws.column_dimensions["C"].width=45; ws.column_dimensions["G"].width=20
    ws.freeze_panes="A4"; ws.auto_filter.ref=ws.dimensions


def _hoja_hoy(wb, tareas_hoy, hoy_str):
    ws = wb.create_sheet(title="Hoy"); ws.sheet_properties.tabColor=_BLUE; ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:G1"); c=ws["A1"]
    c.value=f"  TAREAS DEL DÍA — {datetime.strptime(hoy_str,'%Y-%m-%d').strftime('%d / %m / %Y')}"
    c.fill=_xf("0D2040"); c.font=_xfn(bold=True,size=13,color=_BLUE)
    c.alignment=_xal(h="left"); c.border=_xb(); ws.row_dimensions[1].height=32
    th=len(tareas_hoy); ch=sum(1 for t in tareas_hoy if t["completada"]==1)
    ws.merge_cells("A2:G2"); c=ws["A2"]
    c.value=f"  {th} tarea{'s' if th!=1 else ''} programadas hoy  ·  {ch} completadas  ·  {th-ch} pendientes"
    c.fill=_xf(_DARK2); c.font=_xfn(color=_TXT_M,size=9,italic=True)
    c.alignment=_xal(h="left"); c.border=_xb(); ws.row_dimensions[2].height=18
    for ci,h in enumerate(["ID","Código","Descripción","Categoría","Estado","Usuario","Notas"],start=1):
        _xhdr(ws.cell(row=3,column=ci,value=h),bg="0D2040",fg=_BLUE)
    ws.row_dimensions[3].height=20
    if not tareas_hoy:
        ws.merge_cells("A4:G4"); e=ws["A4"]; e.value="No hay tareas programadas para hoy."
        e.fill=_xf(_DARK2); e.font=_xfn(color=_TXT_M,italic=True); e.alignment=_xal(); e.border=_xb()
    else:
        for ri,t in enumerate(tareas_hoy,start=4):
            done=t["completada"]==1; est="✔  Completada" if done else "●  Pendiente"
            ec=_ACCENT if done else _AMBER; rbg="1A4A24" if done else "3D2E00"
            vals=[t["id"],t["codigo"] or "—",t["descripcion"],t["categoria"] or "General",est,t.get("username") or "—",""]
            for ci,val in enumerate(vals,start=1):
                c=ws.cell(row=ri,column=ci,value=val)
                c.fill=_xf(rbg if ci==5 else _DARK2)
                c.font=_xfn(bold=(ci==5),color=ec if ci==5 else (_TXT_M if ci in (1,6) else _TXT_W),size=9 if ci in (1,6) else 10)
                c.alignment=_xal(h="center" if ci in (1,5) else "left",wrap=(ci==3)); c.border=_xb()
            ws.row_dimensions[ri].height=16
    _xaw(ws); ws.column_dimensions["C"].width=45; ws.column_dimensions["G"].width=20
    ws.freeze_panes="A4"; ws.auto_filter.ref=ws.dimensions


@app.route("/exportar")
@login_required
def exportar():
    hoy      = datetime.now().strftime("%Y-%m-%d")
    user_id  = session.get("user_id")
    admin    = session.get("es_admin") == 1
    username = session.get("user", "sistema")

    conn = get_connection()
    base_q = """
        SELECT t.id, t.descripcion, t.categoria, t.fecha,
               t.completada, t.codigo, t.usuario_id, u.username
        FROM tareas t LEFT JOIN usuarios u ON t.usuario_id = u.id
    """
    with conn.cursor() as cursor:
        if admin:
            cursor.execute(base_q + " ORDER BY t.categoria, t.id")
            todas = cursor.fetchall()
            cursor.execute(base_q + " WHERE t.fecha=%s ORDER BY t.id", (hoy,))
            tareas_hoy = cursor.fetchall()
        else:
            cursor.execute(base_q + " WHERE t.usuario_id=%s ORDER BY t.categoria, t.id", (user_id,))
            todas = cursor.fetchall()
            cursor.execute(base_q + " WHERE t.usuario_id=%s AND t.fecha=%s ORDER BY t.id", (user_id, hoy))
            tareas_hoy = cursor.fetchall()
    conn.close()

    cats_dict = defaultdict(list)
    for t in todas:
        cats_dict[t["categoria"] or "General"].append(t)

    wb = Workbook()
    _hoja_resumen(wb, todas, hoy, admin, username)
    for idx, (cat_nombre, tareas_cat) in enumerate(sorted(cats_dict.items())):
        _hoja_categoria(wb, cat_nombre, tareas_cat, idx)
    _hoja_hoy(wb, tareas_hoy, hoy)

    buffer = io.BytesIO()
    wb.save(buffer); buffer.seek(0)
    return send_file(
        buffer, as_attachment=True,
        download_name=f"Reporte_Tareas_{hoy}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ── Inicialización (se ejecuta con gunicorn Y con python directo) ─────────
inicializar_todo()
inicializar_formacion()

if __name__ == "__main__":
    app.run(debug=os.getenv("FLASK_DEBUG","0")=="1", host="0.0.0.0", port=int(os.getenv("PORT", 5000)))