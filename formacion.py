"""
formacion.py  —  Módulo de Formación para el Gestor de Tareas
Base de datos: MariaDB (misma BD que app_web.py)
"""

from flask import Blueprint, render_template, request, redirect, session, url_for, jsonify, send_file, abort, current_app
from functools import wraps
from datetime import datetime
import psycopg2
import psycopg2.extras
import openpyxl
import io
import os
import unicodedata
from dotenv import load_dotenv

load_dotenv()

# ── Blueprint ──────────────────────────────────────────────────────────────────
formacion_bp = Blueprint(
    "formacion", __name__,
    template_folder="templates"
)

# ── Decorador de login ─────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

# ── Conexión ───────────────────────────────────────────────────────────────────
def get_form_conn():
    import os
    return psycopg2.connect(
        host=os.getenv("DB_HOST", "localhost"),
        port=int(os.getenv("DB_PORT", 5432)),
        dbname=os.getenv("DB_NAME", "gestor_tareas"),
        user=os.getenv("DB_USER", "postgres"),
        password=os.getenv("DB_PASSWORD", ""),
        cursor_factory=psycopg2.extras.RealDictCursor,
    )
# ── Inicialización de tablas ───────────────────────────────────────────────────
def inicializar_formacion():
    conn = get_form_conn()
    with conn.cursor() as cursor:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS alumnos (
                id           INT           NOT NULL AUTO_INCREMENT,
                curso        VARCHAR(200),
                nombre       VARCHAR(200)  NOT NULL,
                progreso     NUMERIC(5,2)  NOT NULL DEFAULT 0.00,
                examenes     INT           NOT NULL DEFAULT 0,
                fecha_inicio DATE,
                fecha_fin    DATE,
                supera_75    TINYINT(1)    NOT NULL DEFAULT 0,
                telefono     VARCHAR(30),
                tutor_id     INT,
                archivado    TINYINT(1)    NOT NULL DEFAULT 0,
                archivado_at DATETIME,
                created_at   DATETIME      NOT NULL DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (id)
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS historial_snapshots (
                id           INT          NOT NULL AUTO_INCREMENT,
                tutor_id     INT,
                fecha        VARCHAR(20),
                label        VARCHAR(200),
                total        INT,
                superan_75   INT,
                pct_exito    NUMERIC(5,2),
                avg_progreso NUMERIC(5,2),
                created_at   DATETIME     NOT NULL DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (id)
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS historial_automatico (
                id             SERIAL NOT NULL,
                tutor_id       INT,
                fecha          VARCHAR(20),
                evento         VARCHAR(300),
                total_alumnos  INT,
                total_cursos   INT,
                created_at     DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (id)
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS alarmas_completadas (
                id         INT         NOT NULL AUTO_INCREMENT,
                tutor_id   INT         NOT NULL,
                clave      VARCHAR(200) NOT NULL,
                fecha_dia  DATE        NOT NULL,
                created_at DATETIME    NOT NULL DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (id),
                UNIQUE KEY uq_alarma (tutor_id, clave, fecha_dia)
            )
        """)
    conn.commit()
    conn.close()
    print("✅ Tablas de formación inicializadas en MariaDB.")

# ── Helpers ────────────────────────────────────────────────────────────────────
def _registrar_evento_historico(tutor_id, evento, conn):
    with conn.cursor() as cursor:
        cursor.execute(
            "SELECT COUNT(*) AS total, COUNT(DISTINCT curso) AS cursos FROM alumnos WHERE tutor_id=%s",
            (tutor_id,)
        )
        row = cursor.fetchone()
        total_alumnos = row["total"] if row else 0
        total_cursos  = row["cursos"] if row else 0
        fecha = datetime.now().strftime("%Y-%m-%d %H:%M")
        cursor.execute("""
            INSERT INTO historial_automatico (tutor_id, fecha, evento, total_alumnos, total_cursos)
            VALUES (%s, %s, %s, %s, %s)
        """, (tutor_id, fecha, evento, total_alumnos, total_cursos))

def _safe_float(val):
    try:
        return float(str(val).replace("%", "").replace(",", ".").strip())
    except (ValueError, TypeError):
        return 0.0

def _safe_int(val):
    try:
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return 0

def _safe_date(val):
    if val is None:
        return None
    from datetime import datetime as _dt, date
    if isinstance(val, (_dt, date)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return _dt.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s if s else None


# ── Motor de alarmas ──────────────────────────────────────────────────────────
def _generar_alarmas(tutor_id):
    from datetime import date as _date_cls
    import urllib.parse

    hoy = _date_cls.today()
    conn = get_form_conn()
    with conn.cursor() as cursor:
        cursor.execute(
            "SELECT * FROM alumnos WHERE tutor_id=%s AND (archivado IS NULL OR archivado=0) ORDER BY fecha_fin ASC, progreso ASC",
            (tutor_id,)
        )
        alumnos = cursor.fetchall()
    conn.close()

    alarmas = []
    cursos_foro_inicio = set()
    cursos_foro_cierre = set()

    def _dias_restantes(fecha_fin_str):
        if not fecha_fin_str: return None
        try:
            return (_date_cls.fromisoformat(str(fecha_fin_str)[:10]) - hoy).days
        except Exception:
            return None

    def _dias_desde_inicio(fecha_inicio_str):
        if not fecha_inicio_str: return None
        try:
            return (hoy - _date_cls.fromisoformat(str(fecha_inicio_str)[:10])).days
        except Exception:
            return None

    def _wa_link(telefono, mensaje):
        if not telefono: return None
        tel = str(telefono).strip().replace(" ", "").replace("-", "").replace("+", "")
        return f"https://wa.me/{tel}?text={urllib.parse.quote(mensaje)}"

    for a in alumnos:
        alumno_id    = a["id"]
        nombre       = a.get("nombre", "")
        curso        = a.get("curso") or "Sin curso"
        progreso     = float(a.get("progreso") or 0)
        supera_75    = int(a.get("supera_75") or 0)
        telefono     = a.get("telefono") or ""
        fecha_fin    = a.get("fecha_fin")
        fecha_inicio = a.get("fecha_inicio")
        dias_r       = _dias_restantes(fecha_fin)
        dias_i       = _dias_desde_inicio(fecha_inicio)

        # 1. CURSO VENCIDO Y NO APROBADO
        if dias_r is not None and dias_r < 0 and not supera_75:
            msg = (f"Hola {nombre} 👋, te contactamos porque tu curso *{curso}* ha finalizado "
                   f"y tu progreso está en {progreso:.0f}%. ¿Podemos ayudarte a completarlo? ¡Estamos aquí para acompañarte!")
            alarmas.append({"clave": f"vencido:{alumno_id}", "tipo": "vencido_bajo", "prioridad": 1,
                "emoji": "🔴", "titulo": "Curso vencido sin aprobar",
                "descripcion": f"El curso finalizó hace {abs(dias_r)} día{'s' if abs(dias_r)!=1 else ''} y el alumno tiene {progreso:.0f}% de progreso.",
                "accion": "Contactar urgente por WhatsApp y evaluar prórroga.",
                "alumno_id": alumno_id, "alumno_nombre": nombre, "curso": curso,
                "telefono": telefono, "dias_restantes": dias_r, "wa_link": _wa_link(telefono, msg)})

        # 2. PROGRESO CRÍTICO CON POCO TIEMPO
        if dias_r is not None and 0 <= dias_r <= 20 and progreso < 25 and not supera_75:
            msg = (f"Hola {nombre} 👋, notamos que tu progreso en *{curso}* es de solo {progreso:.0f}% "
                   f"y quedan {dias_r} días para finalizar. ¡Te acompañamos para que puedas alcanzar el objetivo! ¿Cuándo podemos hablar?")
            alarmas.append({"clave": f"critico:{alumno_id}", "tipo": "progreso_critico", "prioridad": 1,
                "emoji": "🔴", "titulo": "Alumno en riesgo crítico",
                "descripcion": f"Progreso {progreso:.0f}% con solo {dias_r} días restantes. Riesgo alto de no aprobar.",
                "accion": "Intervención inmediata. Llamar o enviar WA con plan de acción.",
                "alumno_id": alumno_id, "alumno_nombre": nombre, "curso": curso,
                "telefono": telefono, "dias_restantes": dias_r, "wa_link": _wa_link(telefono, msg)})

        # 3. FECHA FIN EN 7 DÍAS
        elif dias_r is not None and 1 <= dias_r <= 7 and not supera_75:
            msg = (f"Hola {nombre} 👋, te recordamos que tu curso *{curso}* finaliza en {dias_r} día{'s' if dias_r!=1 else ''}. "
                   f"Tu progreso actual es {progreso:.0f}%. ¡Ánimo, todavía estás a tiempo! 💪")
            alarmas.append({"clave": f"fin7:{alumno_id}", "tipo": "fin_7dias", "prioridad": 1,
                "emoji": "🔴", "titulo": f"Cierre en {dias_r} día{'s' if dias_r!=1 else ''} — sin aprobar",
                "descripcion": f"El curso termina el {fecha_fin}. Progreso: {progreso:.0f}%.",
                "accion": "Enviar WA de recordatorio urgente de cierre.",
                "alumno_id": alumno_id, "alumno_nombre": nombre, "curso": curso,
                "telefono": telefono, "dias_restantes": dias_r, "wa_link": _wa_link(telefono, msg)})

        # 4. FECHA FIN EN 8–14 DÍAS
        elif dias_r is not None and 8 <= dias_r <= 14 and not supera_75:
            msg = (f"Hola {nombre} 👋, quedan {dias_r} días para que finalice tu curso *{curso}*. "
                   f"Llevas un progreso de {progreso:.0f}%. ¡Te animamos a avanzar para alcanzar el objetivo! 🎓")
            alarmas.append({"clave": f"fin14:{alumno_id}", "tipo": "fin_14dias", "prioridad": 2,
                "emoji": "🟡", "titulo": f"Cierre en {dias_r} días — recordatorio",
                "descripcion": f"Quedan 2 semanas. Progreso: {progreso:.0f}%. Buen momento para hacer seguimiento.",
                "accion": "Enviar WA motivacional con recordatorio de fechas.",
                "alumno_id": alumno_id, "alumno_nombre": nombre, "curso": curso,
                "telefono": telefono, "dias_restantes": dias_r, "wa_link": _wa_link(telefono, msg)})

        # 5. FECHA FIN EN 15–30 DÍAS Y PROGRESO BAJO
        elif dias_r is not None and 15 <= dias_r <= 30 and progreso < 50 and not supera_75:
            msg = (f"Hola {nombre} 👋, ¿cómo vas con tu curso *{curso}*? Tienes {dias_r} días para completarlo "
                   f"y tu progreso es {progreso:.0f}%. Planifiquemos juntos para que llegues a la meta 🎯")
            alarmas.append({"clave": f"fin30:{alumno_id}", "tipo": "fin_30dias", "prioridad": 2,
                "emoji": "🟡", "titulo": f"Progreso bajo con cierre en {dias_r} días",
                "descripcion": f"Progreso {progreso:.0f}% — menos del 50% con menos de 30 días disponibles.",
                "accion": "Planificar seguimiento y enviar WA con agenda de trabajo.",
                "alumno_id": alumno_id, "alumno_nombre": nombre, "curso": curso,
                "telefono": telefono, "dias_restantes": dias_r, "wa_link": _wa_link(telefono, msg)})

        # 6. INICIO HOY
        if dias_i is not None and dias_i == 0:
            msg = (f"¡Bienvenido/a {nombre}! 🎉 Hoy comenzás tu curso *{curso}*. "
                   f"Estamos muy contentos de acompañarte. ¡Cualquier consulta, aquí estamos! 😊")
            alarmas.append({"clave": f"inicio_hoy:{alumno_id}", "tipo": "inicio_hoy", "prioridad": 2,
                "emoji": "🟢", "titulo": "Inicio de curso HOY",
                "descripcion": f"El alumno empieza hoy el curso {curso}.",
                "accion": "Enviar WA de bienvenida y verificar acceso al aula virtual.",
                "alumno_id": alumno_id, "alumno_nombre": nombre, "curso": curso,
                "telefono": telefono, "dias_restantes": dias_r, "wa_link": _wa_link(telefono, msg)})
            if curso not in cursos_foro_inicio:
                cursos_foro_inicio.add(curso)
                alarmas.append({"clave": f"foro_inicio:{curso}", "tipo": "foro_inicio", "prioridad": 2,
                    "emoji": "🟢", "titulo": f"Publicar foro de inicio — {curso}",
                    "descripcion": f"El curso {curso} arranca hoy. Publicar el foro de presentación e inicio de la primera unidad.",
                    "accion": "Abrir hilo de presentación y foro de inicio de unidad en el aula virtual.",
                    "alumno_id": None, "alumno_nombre": f"Curso completo: {curso}", "curso": curso,
                    "telefono": "", "dias_restantes": None, "wa_link": None})

        # 7. INICIO HACE 1–3 DÍAS
        elif dias_i is not None and 1 <= dias_i <= 3:
            alarmas.append({"clave": f"inicio_reciente:{alumno_id}", "tipo": "inicio_reciente", "prioridad": 3,
                "emoji": "🔵", "titulo": f"Inicio reciente ({dias_i} día{'s' if dias_i!=1 else ''} atrás)",
                "descripcion": f"El alumno inició el {fecha_inicio}. Verificar que ingresó correctamente al aula.",
                "accion": "Confirmar acceso al aula virtual y progreso inicial.",
                "alumno_id": alumno_id, "alumno_nombre": nombre, "curso": curso,
                "telefono": telefono, "dias_restantes": dias_r, "wa_link": None})

        # 8. FORO DE CIERRE DE UNIDAD
        if dias_r is not None and 3 <= dias_r <= 7 and curso not in cursos_foro_cierre:
            cursos_foro_cierre.add(curso)
            alarmas.append({"clave": f"foro_cierre:{curso}", "tipo": "foro_cierre", "prioridad": 2,
                "emoji": "🟡", "titulo": f"Publicar foro de cierre — {curso}",
                "descripcion": f"El curso {curso} cierra en {dias_r} días. Publicar foro de cierre.",
                "accion": "Abrir foro de cierre, reflexión final y recordatorio de evaluación.",
                "alumno_id": None, "alumno_nombre": f"Curso completo: {curso}", "curso": curso,
                "telefono": "", "dias_restantes": dias_r, "wa_link": None})

        # 9. SIN TELÉFONO REGISTRADO
        if not telefono and ((dias_r is not None and dias_r <= 30) or progreso < 50):
            alarmas.append({"clave": f"sin_tel:{alumno_id}", "tipo": "sin_telefono", "prioridad": 3,
                "emoji": "🔵", "titulo": "Sin teléfono de contacto",
                "descripcion": f"El alumno no tiene teléfono registrado y requiere seguimiento (progreso {progreso:.0f}%).",
                "accion": "Registrar número de WhatsApp en la ficha del alumno.",
                "alumno_id": alumno_id, "alumno_nombre": nombre, "curso": curso,
                "telefono": "", "dias_restantes": dias_r, "wa_link": None})

        # 10. PROGRESO MENOR AL 40%
        if progreso < 40 and not supera_75 and not (
            (dias_r is not None and dias_r <= 20 and progreso < 25) or
            (dias_r is not None and dias_r < 0)
        ):
            msg = (f"Hola {nombre} 👋, notamos que tu progreso en *{curso}* es de {progreso:.0f}%, "
                   f"que está por debajo del 40% esperado. "
                   f"¿Hay algo en lo que podamos ayudarte para avanzar? ¡Estamos aquí para acompañarte! 💪")
            alarmas.append({"clave": f"bajo40:{alumno_id}", "tipo": "progreso_bajo_40", "prioridad": 2,
                "emoji": "🟠", "titulo": f"Progreso bajo el 40% — {progreso:.0f}%",
                "descripcion": f"El alumno lleva {progreso:.0f}% de avance, por debajo del umbral mínimo del 40%."
                               + (f" Quedan {dias_r} día{'s' if dias_r != 1 else ''}." if dias_r is not None and dias_r > 0 else ""),
                "accion": "Contactar al alumno para identificar obstáculos y reforzar el acompañamiento.",
                "alumno_id": alumno_id, "alumno_nombre": nombre, "curso": curso,
                "telefono": telefono, "dias_restantes": dias_r, "wa_link": _wa_link(telefono, msg)})

    alarmas.sort(key=lambda x: (x["prioridad"], x["dias_restantes"] if x["dias_restantes"] is not None else 9999))
    return alarmas


def _get_completadas_hoy(tutor_id):
    from datetime import date as _d
    hoy = _d.today().isoformat()
    conn = get_form_conn()
    with conn.cursor() as cursor:
        cursor.execute(
            "SELECT clave FROM alarmas_completadas WHERE tutor_id=%s AND fecha_dia=%s",
            (tutor_id, hoy)
        )
        rows = cursor.fetchall()
    conn.close()
    return {r["clave"] for r in rows}


# ── Ruta: listado de alumnos + carga de Excel ──────────────────────────────────
@formacion_bp.route("/formacion", methods=["GET", "POST"])
@login_required
def formacion():
    tutor_id = session.get("user_id")
    errores  = []
    exito    = None

    if request.method == "POST":
        archivo = request.files.get("excel")
        if not archivo or not archivo.filename.endswith((".xlsx", ".xls")):
            errores.append("Por favor sube un archivo Excel válido (.xlsx o .xls).")
        else:
            try:
                wb = openpyxl.load_workbook(io.BytesIO(archivo.read()), data_only=True)
                ws = wb.active

                headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

                def norm(s):
                    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
                headers_norm = [norm(h) for h in headers]

                def col_idx(posibles):
                    for p in posibles:
                        pn = norm(p)
                        for i, h in enumerate(headers_norm):
                            if pn in h: return i
                    return None

                def col_idx_excl(posibles, excluir):
                    for p in posibles:
                        pn = norm(p)
                        for i, h in enumerate(headers_norm):
                            if pn in h and i != excluir: return i
                    return None

                idx_curso    = col_idx(["del curso", "curso", "materia", "asignatura"])
                idx_nombre   = col_idx_excl(["nombre", "alumno", "estudiante"], idx_curso)
                idx_progreso = col_idx(["progreso", "avance", "progress"])
                idx_examenes = col_idx(["examen", "exam", "evaluac", "prueba"])
                idx_inicio   = col_idx(["fecha inicio", "fecha de inicio", "f. inicio", "inicio"])
                idx_fin      = None
                for i, h in enumerate(headers_norm):
                    if "fin" in h and "inicio" not in h:
                        idx_fin = i; break
                idx_telefono = col_idx(["telefono", "phone", "celular", "whatsapp"])

                if idx_nombre is None:
                    errores.append("No se encontró columna de Nombre. Revisa el Excel.")
                else:
                    conn  = get_form_conn()
                    count = 0
                    with conn.cursor() as cursor:
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if not any(row): continue
                            nombre = str(row[idx_nombre]).strip() if idx_nombre is not None and row[idx_nombre] else None
                            if not nombre or nombre.lower() in ("none", "nan", ""): continue

                            progreso  = _safe_float(row[idx_progreso]) if idx_progreso is not None and idx_progreso < len(row) else 0.0
                            examenes  = _safe_int(row[idx_examenes])   if idx_examenes is not None and idx_examenes < len(row) else 0
                            f_inicio  = _safe_date(row[idx_inicio])    if idx_inicio   is not None and idx_inicio   < len(row) else None
                            f_fin     = _safe_date(row[idx_fin])       if idx_fin      is not None and idx_fin      < len(row) else None
                            telefono  = str(row[idx_telefono]).strip() if idx_telefono is not None and idx_telefono < len(row) and row[idx_telefono] else None
                            curso     = str(row[idx_curso]).strip()    if idx_curso    is not None and idx_curso    < len(row) and row[idx_curso]    else None
                            supera_75 = 1 if progreso >= 75 else 0

                            cursor.execute("""
                                INSERT INTO alumnos (curso, nombre, progreso, examenes, fecha_inicio, fecha_fin,
                                                     supera_75, telefono, tutor_id)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                            """, (curso, nombre, progreso, examenes, f_inicio, f_fin, supera_75, telefono, tutor_id))
                            count += 1

                        _registrar_evento_historico(tutor_id, f"Importación (+{count} alumnos)", conn)
                    conn.commit()
                    conn.close()
                    exito = f"✅ {count} alumnos importados correctamente."

            except Exception as e:
                errores.append(f"Error al procesar el archivo: {e}")

    conn = get_form_conn()
    with conn.cursor() as cursor:
        cursor.execute(
            "SELECT * FROM alumnos WHERE tutor_id=%s AND (archivado IS NULL OR archivado=0) ORDER BY id DESC",
            (tutor_id,)
        )
        alumnos = cursor.fetchall()
        cursor.execute(
            "SELECT COUNT(DISTINCT curso) AS n FROM alumnos WHERE tutor_id=%s AND archivado=1",
            (tutor_id,)
        )
        row_arch = cursor.fetchone()
        archivados_count = row_arch["n"] if row_arch else 0
    conn.close()

    alarmas_hoy        = _generar_alarmas(tutor_id)
    completadas_hoy    = _get_completadas_hoy(tutor_id)
    alarmas_pendientes = sum(1 for a in alarmas_hoy if a["clave"] not in completadas_hoy)

    return render_template("formacion.html", alumnos=alumnos, errores=errores, exito=exito,
                           alarmas_pendientes=alarmas_pendientes, archivados_count=archivados_count)


# ── Ruta: editar alumno (teléfono) ─────────────────────────────────────────────
@formacion_bp.route("/formacion/editar/<int:alumno_id>", methods=["POST"])
@login_required
def editar_alumno(alumno_id):
    telefono = request.form.get("telefono", "").strip()
    conn = get_form_conn()
    with conn.cursor() as cursor:
        cursor.execute("UPDATE alumnos SET telefono=%s WHERE id=%s AND tutor_id=%s",
                       (telefono, alumno_id, session["user_id"]))
    conn.commit()
    conn.close()
    return redirect(url_for("formacion.formacion"))


# ── Ruta: eliminar alumno ──────────────────────────────────────────────────────
@formacion_bp.route("/formacion/eliminar/<int:alumno_id>")
@login_required
def eliminar_alumno(alumno_id):
    conn = get_form_conn()
    with conn.cursor() as cursor:
        cursor.execute("DELETE FROM alumnos WHERE id=%s AND tutor_id=%s",
                       (alumno_id, session["user_id"]))
    _registrar_evento_historico(session["user_id"], "Eliminación de alumno", conn)
    conn.commit()
    conn.close()
    return redirect(url_for("formacion.formacion"))


# ── Ruta: borrar TODOS los alumnos ────────────────────────────────────────────
@formacion_bp.route("/formacion/borrar_todos", methods=["POST"])
@login_required
def borrar_todos():
    conn = get_form_conn()
    with conn.cursor() as cursor:
        cursor.execute("DELETE FROM alumnos WHERE tutor_id=%s", (session["user_id"],))
    _registrar_evento_historico(session["user_id"], "Borrado total", conn)
    conn.commit()
    conn.close()
    return redirect(url_for("formacion.formacion"))


# ── Ruta: archivar curso completo ─────────────────────────────────────────────
@formacion_bp.route("/formacion/archivar_curso", methods=["POST"])
@login_required
def archivar_curso():
    tutor_id = session["user_id"]
    curso    = request.form.get("curso", "").strip()
    if not curso:
        return redirect(url_for("formacion.formacion"))
    ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn  = get_form_conn()
    with conn.cursor() as cursor:
        cursor.execute("""
            UPDATE alumnos SET archivado=1, archivado_at=%s
            WHERE tutor_id=%s AND curso=%s AND (archivado IS NULL OR archivado=0)
        """, (ahora, tutor_id, curso))
    _registrar_evento_historico(tutor_id, f"Curso archivado: {curso}", conn)
    conn.commit()
    conn.close()
    return redirect(url_for("formacion.formacion"))


# ── Ruta: restaurar curso archivado ───────────────────────────────────────────
@formacion_bp.route("/formacion/restaurar_curso", methods=["POST"])
@login_required
def restaurar_curso():
    tutor_id = session["user_id"]
    curso    = request.form.get("curso", "").strip()
    if not curso:
        return redirect(url_for("formacion.formacion_archivados"))
    conn = get_form_conn()
    with conn.cursor() as cursor:
        cursor.execute("""
            UPDATE alumnos SET archivado=0, archivado_at=NULL
            WHERE tutor_id=%s AND curso=%s AND archivado=1
        """, (tutor_id, curso))
    _registrar_evento_historico(tutor_id, f"Curso restaurado: {curso}", conn)
    conn.commit()
    conn.close()
    return redirect(url_for("formacion.formacion_archivados"))


# ── Ruta: vista de archivados ──────────────────────────────────────────────────
@formacion_bp.route("/formacion/archivados")
@login_required
def formacion_archivados():
    tutor_id = session["user_id"]
    conn     = get_form_conn()
    with conn.cursor() as cursor:
        cursor.execute(
            "SELECT * FROM alumnos WHERE tutor_id=%s AND archivado=1 ORDER BY archivado_at DESC, curso, nombre",
            (tutor_id,)
        )
        alumnos_arch = cursor.fetchall()
    conn.close()

    for a in alumnos_arch:
        a["progreso"]  = float(a.get("progreso") or 0)
        a["examenes"]  = int(a.get("examenes") or 0)
        a["supera_75"] = int(a.get("supera_75") or 0)
        a["curso"]     = a.get("curso") or "Sin curso"

    from collections import defaultdict
    grupos_raw = defaultdict(list)
    for a in alumnos_arch:
        grupos_raw[a["curso"]].append(a)

    def _fecha_archivo_grupo(alumnos_lista):
        fechas = [str(a.get("archivado_at") or "") for a in alumnos_lista]
        return max(fechas) if fechas else ""

    grupos_ordenados = sorted(grupos_raw.items(),
                              key=lambda kv: _fecha_archivo_grupo(kv[1]), reverse=True)

    grupos = []
    for curso_nombre, lista in grupos_ordenados:
        total     = len(lista)
        superan   = sum(1 for a in lista if a["supera_75"] == 1)
        pct       = round(superan / total * 100, 1) if total else 0
        avg_prog  = round(sum(a["progreso"] for a in lista) / total, 1) if total else 0
        examenes  = sum(a["examenes"] for a in lista)
        fecha_arch = (str(lista[0].get("archivado_at") or ""))[:10]
        fecha_ini  = min((str(a.get("fecha_inicio") or "9999") for a in lista if a.get("fecha_inicio")), default="—")
        fecha_fin  = max((str(a.get("fecha_fin")    or "0000") for a in lista if a.get("fecha_fin")),    default="—")
        grupos.append({
            "curso": curso_nombre, "alumnos": lista, "total": total,
            "superan": superan, "pct_exito": pct, "avg_progreso": avg_prog,
            "total_examenes": examenes, "fecha_archivo": fecha_arch,
            "fecha_inicio": fecha_ini if fecha_ini != "9999" else "—",
            "fecha_fin":    fecha_fin if fecha_fin != "0000" else "—",
        })

    return render_template("formacion_archivados.html",
                           grupos=grupos,
                           total_archivados=len(alumnos_arch),
                           total_cursos_arch=len(grupos))


# ── Ruta: guardar snapshot ────────────────────────────────────────────────────
@formacion_bp.route("/formacion/guardar_snapshot", methods=["POST"])
@login_required
def guardar_snapshot():
    tutor_id = session["user_id"]
    label    = request.form.get("label", "").strip() or datetime.now().strftime("%d/%m/%Y")
    fecha    = datetime.now().strftime("%Y-%m-%d")

    conn = get_form_conn()
    with conn.cursor() as cursor:
        cursor.execute(
            "SELECT * FROM alumnos WHERE tutor_id=%s AND (archivado IS NULL OR archivado=0)", (tutor_id,)
        )
        alumnos = cursor.fetchall()
        total        = len(alumnos)
        superan_75   = sum(1 for a in alumnos if a["supera_75"] == 1)
        pct_exito    = round(superan_75 / total * 100, 1) if total else 0
        avg_progreso = round(sum(float(a["progreso"] or 0) for a in alumnos) / total, 1) if total else 0
        cursor.execute("""
            INSERT INTO historial_snapshots (tutor_id, fecha, label, total, superan_75, pct_exito, avg_progreso)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (tutor_id, fecha, label, total, superan_75, pct_exito, avg_progreso))
    conn.commit()
    conn.close()
    return redirect(url_for("formacion.formacion_dashboard"))


# ── Ruta: borrar snapshot ──────────────────────────────────────────────────────
@formacion_bp.route("/formacion/borrar_snapshot/<int:snap_id>")
@login_required
def borrar_snapshot(snap_id):
    conn = get_form_conn()
    with conn.cursor() as cursor:
        cursor.execute("DELETE FROM historial_snapshots WHERE id=%s AND tutor_id=%s",
                       (snap_id, session["user_id"]))
    conn.commit()
    conn.close()
    return redirect(url_for("formacion.formacion_dashboard"))


# ── Ruta: borrar historial automático ─────────────────────────────────────────
@formacion_bp.route("/formacion/borrar_historial_auto", methods=["POST"])
@login_required
def borrar_historial_auto():
    conn = get_form_conn()
    with conn.cursor() as cursor:
        cursor.execute("DELETE FROM historial_automatico WHERE tutor_id=%s", (session["user_id"],))
    conn.commit()
    conn.close()
    return redirect(url_for("formacion.formacion_dashboard"))


# ── Ruta: alarmas ─────────────────────────────────────────────────────────────
@formacion_bp.route("/formacion/alarmas")
@login_required
def formacion_alarmas():
    tutor_id    = session["user_id"]
    alarmas     = _generar_alarmas(tutor_id)
    completadas = _get_completadas_hoy(tutor_id)

    for a in alarmas:
        a["completada"] = a["clave"] in completadas

    total      = len(alarmas)
    pendientes = sum(1 for a in alarmas if not a["completada"])
    hechas     = total - pendientes

    from datetime import date as _d
    hoy_str = _d.today().strftime("%A %d de %B de %Y")

    return render_template("formacion_alarmas.html",
                           alarmas=alarmas, total=total,
                           pendientes=pendientes, hechas=hechas, hoy_str=hoy_str)


# ── Ruta: completar / deshacer alarma ─────────────────────────────────────────
@formacion_bp.route("/formacion/alarmas/completar", methods=["POST"])
@login_required
def alarma_completar():
    from datetime import date as _d
    tutor_id = session["user_id"]
    clave    = request.form.get("clave", "").strip()
    accion   = request.form.get("accion", "completar")
    hoy      = _d.today().isoformat()

    if not clave:
        return redirect(url_for("formacion.formacion_alarmas"))

    conn = get_form_conn()
    with conn.cursor() as cursor:
        if accion == "deshacer":
            cursor.execute(
                "DELETE FROM alarmas_completadas WHERE tutor_id=%s AND clave=%s AND fecha_dia=%s",
                (tutor_id, clave, hoy)
            )
        else:
            cursor.execute(
                "INSERT INTO alarmas_completadas (tutor_id, clave, fecha_dia) VALUES (%s,%s,%s)",
                (tutor_id, clave, hoy)
            )
    conn.commit()
    conn.close()
    return redirect(url_for("formacion.formacion_alarmas"))


# ── Ruta: badge JSON ──────────────────────────────────────────────────────────
@formacion_bp.route("/formacion/alarmas/badge")
@login_required
def alarmas_badge():
    tutor_id    = session["user_id"]
    alarmas     = _generar_alarmas(tutor_id)
    completadas = _get_completadas_hoy(tutor_id)
    pendientes  = sum(1 for a in alarmas if a["clave"] not in completadas)
    return jsonify({"pendientes": pendientes})


# ── Ruta: dashboard de formación ───────────────────────────────────────────────
@formacion_bp.route("/formacion/dashboard")
@login_required
def formacion_dashboard():
    tutor_id = session.get("user_id")
    conn     = get_form_conn()
    with conn.cursor() as cursor:
        cursor.execute(
            "SELECT * FROM alumnos WHERE tutor_id=%s AND (archivado IS NULL OR archivado=0) ORDER BY progreso DESC",
            (tutor_id,)
        )
        alumnos = cursor.fetchall()
        cursor.execute(
            "SELECT * FROM historial_snapshots WHERE tutor_id=%s ORDER BY fecha ASC", (tutor_id,)
        )
        snapshots = cursor.fetchall()
        cursor.execute(
            "SELECT * FROM historial_automatico WHERE tutor_id=%s ORDER BY id DESC LIMIT 50", (tutor_id,)
        )
        historial_auto = cursor.fetchall()
    conn.close()

    for a in alumnos:
        a["progreso"]  = float(a.get("progreso") or 0)
        a["examenes"]  = int(a.get("examenes") or 0)
        a["supera_75"] = int(a.get("supera_75") or 0)
        a["curso"]     = a.get("curso") or ""

    cursos         = sorted(set(a["curso"] for a in alumnos if a["curso"]))
    total          = len(alumnos)
    superan_75     = sum(1 for a in alumnos if a["supera_75"] == 1)
    no_superan     = total - superan_75
    pct_exito      = round(superan_75 / total * 100, 1) if total else 0
    avg_progreso   = round(sum(a["progreso"] for a in alumnos) / total, 1) if total else 0
    total_examenes = sum(a["examenes"] for a in alumnos)

    snap_labels = [s["label"]        for s in snapshots]
    snap_pct    = [float(s["pct_exito"] or 0)    for s in snapshots]
    snap_avg    = [float(s["avg_progreso"] or 0) for s in snapshots]
    snap_total  = [s["total"]        for s in snapshots]

    hist_labels  = [h["fecha"]         for h in reversed(historial_auto)]
    hist_alumnos = [h["total_alumnos"] for h in reversed(historial_auto)]
    hist_cursos  = [h["total_cursos"]  for h in reversed(historial_auto)]

    return render_template("formacion_dashboard.html",
                           alumnos=alumnos, cursos=cursos, total=total,
                           superan_75=superan_75, no_superan=no_superan,
                           pct_exito=pct_exito, avg_progreso=avg_progreso,
                           total_examenes=total_examenes,
                           snapshots=snapshots, snap_labels=snap_labels,
                           snap_pct=snap_pct, snap_avg=snap_avg, snap_total=snap_total,
                           historial_auto=historial_auto, hist_labels=hist_labels,
                           hist_alumnos=hist_alumnos, hist_cursos=hist_cursos)


# ── Ruta: descargar modelo de importación ─────────────────────────────────────
@formacion_bp.route("/formacion/descargar_modelo")
@login_required
def descargar_modelo():
    posibles = [
        os.path.join(current_app.root_path, "static", "MODELO_IMPORTAR.xlsx"),
        os.path.join(current_app.root_path, "MODELO_IMPORTAR.xlsx"),
        os.path.join(os.path.dirname(__file__), "MODELO_IMPORTAR.xlsx"),
        os.path.join(os.path.dirname(__file__), "static", "MODELO_IMPORTAR.xlsx"),
    ]
    for ruta in posibles:
        if os.path.exists(ruta):
            return send_file(ruta, as_attachment=True, download_name="MODELO_IMPORTAR.xlsx")
    abort(404, "El archivo MODELO_IMPORTAR.xlsx no se encontró en el servidor.")


# ── Ruta: WhatsApp link ────────────────────────────────────────────────────────
@formacion_bp.route("/formacion/whatsapp/<int:alumno_id>")
@login_required
def whatsapp_alumno(alumno_id):
    conn = get_form_conn()
    with conn.cursor() as cursor:
        cursor.execute("SELECT * FROM alumnos WHERE id=%s", (alumno_id,))
        alumno = cursor.fetchone()
    conn.close()

    if not alumno:
        return redirect(url_for("formacion.formacion"))

    telefono = (alumno["telefono"] or "").strip().replace(" ", "").replace("-", "").replace("+", "")
    if not telefono:
        return redirect(url_for("formacion.formacion"))

    nombre   = alumno["nombre"]
    progreso = alumno["progreso"]
    inicio   = str(alumno["fecha_inicio"] or "—")
    fin      = str(alumno["fecha_fin"]    or "—")

    mensaje = (
        f"Hola {nombre} 👋\n"
        f"Te comparto tu resumen de progreso en el curso:\n\n"
        f"📅 Fecha de inicio: {inicio}\n"
        f"📅 Fecha de fin: {fin}\n"
        f"📊 Progreso actual: {progreso}%\n\n"
        f"{'🎉 ¡Superaste el 75% requerido! Excelente trabajo.' if float(progreso or 0) >= 75 else '⚡ Sigue adelante, ¡puedes lograrlo!'}\n\n"
        f"Cualquier consulta, aquí estoy. ¡Éxitos!"
    )
    import urllib.parse
    url = f"https://wa.me/{telefono}?text={urllib.parse.quote(mensaje)}"
    return redirect(url)


# ── Ruta: exportar alumnos a Excel ────────────────────────────────────────────
@formacion_bp.route("/formacion/exportar_excel")
@login_required
def exportar_excel():
    import io as _io
    from datetime import datetime as _dt, date as _date
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import DataBarRule

    tutor_id = session["user_id"]
    conn     = get_form_conn()
    with conn.cursor() as cursor:
        cursor.execute(
            "SELECT * FROM alumnos WHERE tutor_id=%s AND (archivado IS NULL OR archivado=0) ORDER BY curso, nombre",
            (tutor_id,)
        )
        alumnos = cursor.fetchall()
    conn.close()

    C_DARK="1E3A5F"; C_GREEN="2D9D78"; C_AMBER="D4A017"; C_RED="C0392B"
    C_ALT="F0F4F8"; C_WHITE="FFFFFF"; C_BORDER="CBD5E1"

    def thin():
        s = Side(style="thin", color=C_BORDER)
        return Border(left=s, right=s, top=s, bottom=s)

    def hdr(cell, text, bg=None):
        cell.value = text
        cell.font  = Font(bold=True, color=C_WHITE, name="Arial", size=10)
        cell.fill  = PatternFill("solid", fgColor=bg or C_DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin()

    wb = Workbook()
    ws = wb.active; ws.title = "Alumnos"
    ws.merge_cells("A1:K1"); c = ws["A1"]
    c.value = f"Informe de Formación — {_dt.now().strftime('%d/%m/%Y %H:%M')}  ·  {len(alumnos)} alumnos"
    c.font  = Font(bold=True, size=13, color=C_DARK, name="Arial")
    c.fill  = PatternFill("solid", fgColor="E8F0FA")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26; ws.row_dimensions[2].height = 6

    COLS = ["#","Curso","Nombre","Progreso (%)","Exámenes","Fecha Inicio",
            "Fecha Fin","Supera 75%","Teléfono","Estado","Importado"]
    ws.row_dimensions[3].height = 30
    for c_i, h in enumerate(COLS, 1): hdr(ws.cell(3, c_i), h)

    for r, a in enumerate(alumnos, 4):
        ws.row_dimensions[r].height = 18
        rf = PatternFill("solid", fgColor=C_ALT if r%2==0 else C_WHITE)
        p  = float(a.get("progreso") or 0)
        prog_color = C_GREEN if p >= 75 else (C_AMBER if p >= 50 else C_RED)

        def dc(col, val, fmt=None, bold=False, center=False, color=None):
            cell = ws.cell(r, col, val)
            cell.fill      = rf
            cell.font      = Font(name="Arial", size=9, bold=bold, color=color or "1E293B")
            cell.border    = thin()
            cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
            if fmt: cell.number_format = fmt

        dc(1,  r-3,                                       center=True, color="64748B")
        dc(2,  a.get("curso","—") or "—")
        dc(3,  a.get("nombre",""),                        bold=True)
        dc(4,  p,                                         fmt='0.0"%"', center=True, bold=True, color=prog_color)
        dc(5,  a.get("examenes",0) or 0,                 center=True)
        dc(6,  str(a.get("fecha_inicio","—") or "—"),    center=True)
        dc(7,  str(a.get("fecha_fin","—")    or "—"),    center=True)
        dc(8,  "✔ Sí" if a.get("supera_75") else "✖ No", center=True, bold=True,
           color=C_GREEN if a.get("supera_75") else C_AMBER)
        dc(9,  a.get("telefono","—") or "—")
        dc(10, "✅ Supera 75%" if a.get("supera_75") else "⚠ Bajo 75%",
           center=True, bold=True, color=C_GREEN if a.get("supera_75") else C_AMBER)
        dc(11, (str(a.get("created_at","") or ""))[:10],  center=True, color="64748B")

    last_data = 3 + len(alumnos)
    if alumnos:
        ws.conditional_formatting.add(
            f"D4:D{last_data}",
            DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100, color="2D9D78")
        )
    ws.auto_filter.ref = f"A3:{get_column_letter(len(COLS))}{last_data}"
    ws.freeze_panes    = "A4"
    for i, w in enumerate([5,32,22,14,10,14,14,12,18,16,18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Hoja 2 — Resumen por curso (simplificada) ──
    from collections import defaultdict
    ws2 = wb.create_sheet("Resumen por Curso")
    ws2.merge_cells("A1:H1"); c2 = ws2["A1"]
    c2.value = "Resumen de rendimiento por curso"
    c2.font  = Font(bold=True, size=13, color=C_DARK, name="Arial")
    c2.fill  = PatternFill("solid", fgColor="E8F0FA")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26; ws2.row_dimensions[2].height = 6
    COLS2 = ["Curso","Total Alumnos","Superan 75%","Bajo 75%",
             "Tasa Éxito (%)","Prog. Promedio (%)","Total Exámenes","Prom. Exámenes"]
    ws2.row_dimensions[3].height = 30
    for c_i, h in enumerate(COLS2, 1): hdr(ws2.cell(3, c_i), h)
    resumen = defaultdict(lambda: {"total":0,"superan":0,"prog":0,"exam":0})
    for a in alumnos:
        k = a.get("curso") or "Sin curso"
        resumen[k]["total"]  += 1
        resumen[k]["superan"] += int(a.get("supera_75") or 0)
        resumen[k]["prog"]   += float(a.get("progreso") or 0)
        resumen[k]["exam"]   += int(a.get("examenes") or 0)
    for row_i, (curso, d) in enumerate(sorted(resumen.items()), 4):
        n   = d["total"]; sup = d["superan"]
        pct = round(sup/n*100, 1) if n else 0
        avg = round(d["prog"]/n, 1) if n else 0
        avg_e = round(d["exam"]/n, 1) if n else 0
        rf2 = PatternFill("solid", fgColor=C_ALT if row_i%2==0 else C_WHITE)
        for c_i, (val, fmt, center, bold, color) in enumerate([
            (curso, None, False, True, "1E293B"),(n, None, True, False, "1E293B"),
            (sup, None, True, True, C_GREEN),(n-sup, None, True, False, C_AMBER),
            (pct, '0.0"%"', True, True, C_GREEN if pct>=75 else (C_AMBER if pct>=50 else C_RED)),
            (avg, '0.0"%"', True, False, "1E293B"),(d["exam"], None, True, False, "1E293B"),
            (avg_e, "0.0", True, False, "1E293B"),
        ], 1):
            cell = ws2.cell(row_i, c_i, val)
            cell.fill = rf2; cell.font = Font(name="Arial", size=9, bold=bold, color=color)
            cell.border = thin(); cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
            if fmt: cell.number_format = fmt
    ws2.freeze_panes = "A4"
    for i, w in enumerate([32,14,14,12,16,18,14,16], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"formacion_{_dt.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)